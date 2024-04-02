from openpyxl import Workbook


# wb = openpyxl.load_workbook('example.xlsx')
#
# sheet = wb.get_sheet_by_name('Sheet1')

wb = Workbook()
ws = wb.active

sheet = wb.create_sheet("Mysheet")
sheet['A1'] = 'Hello world!'

wb.save('mah_sheet.xlsx')
